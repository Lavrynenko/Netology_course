import os #Подгружаем библиотеку для обработки системных команд
import openpyxl #Подгружаем библиотеку для обработки  Excel-файлов
import re

path = os.chdir('c:\\Users\\Oleksiy\\Desktop\\Полный список аптек в Украине\\')

list_directory = os.listdir(path) #В переменную загоняем список файлов в целевой директории
print('Список всех файлов \n', list_directory) #Выводим полный список файлов в каталоге
print(type(list_directory))
for i in list_directory:
    print('Начинаем обработку файла:', i)
    wb = openpyxl.load_workbook(i)
    sheets = wb.get_sheet_names()  # Получили полный список листов в файле и загнали в список
    sheet = wb.get_sheet_by_name(sheets[0])
    row_count = sheet.max_row  # Определили количество строк
    column_count = sheet.max_column  # Определили количество колонок

    print('В документе', i, '\n - строк: ', row_count, '\n - колонок:', column_count)
    print('\n')

    number_row = 2 #Вводим переменную, отвечающую за номер строки
    letter_colomn = 'D' #Вводим переменную, отвечающую за колонку

    while number_row < row_count:
        regxp = '((067|098|050|066|097|068|099|093|096)\d{7})' #Регулярное выражение для поиска мобильных номеров
        number_row = str(number_row)
        row = letter_colomn + number_row #Создаем переменную ячейки, с которой будем работать
        row = str(row) #Конвертируем ее в строковую
        temp_number = sheet[row].value #В переменную загоняем значение ячейки
        temp_number = str(temp_number) #Конвертируем это значение в строковую

        temp_number = temp_number.replace('(', '').replace(')', '').replace(' ', '').replace('-', '').replace('.',
                                                                                                              '').replace(
            ',', '')  #Обрезали все лишнее в номере - тире, точки, скобочки
        result_number = re.findall(regxp, temp_number)  #Начинаем парсить по номеру (напоминаю, номер уже окультурили)

        #typer = type(result_number)
        lens = len(result_number) #Вводим переменную, отвечающую за количество элементов после парсинга
        result_number = str(result_number)
        if lens == 1: #Если элемент один
            number_to_file = result_number[3:12] #Берем все цифры в указанном промежутке
            with open('number.txt', 'a', encoding='utf8') as f: #Создаем файл - если его не было раньше - в режиме дозаписи
                f.write(number_to_file + '\n') #Сохраняем в файл найденный и очищенный номер мобильного телефона
        elif lens == 2: #Если элементов два
            number_to_file = result_number[3:12]  # Берем все цифры в указанном промежутке
            number_to_file2 = result_number[26:35] #Берем все цифры в указанном промежутке
            with open('number.txt', 'a', encoding='utf8') as f: #Создаем файл - если его не было раньше - в режиме дозаписи
                f.write(number_to_file + '\n' + number_to_file2 + '\n') #Сохраняем в файл найденный и очищенный номер мобильного телефона

        number_row = int(number_row) #Конвертируем переменную в инт
        number_row = number_row + 1 #Увеличиваем на единицу - что бы работать с ячейкой ниже