import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Fill
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook
import glob #Библиотека для работы с файлами
import re #Библиотека для парсинга регулярными выражениями

#Создаем итоговый файл
final_wb = openpyxl.Workbook()

print('Начинаем работу: \n')

#Формируем список файлов для работы
original_path = 'c:\\Users\\Oleksiy\\Desktop\\Git\\Netology\\Excel\\Original\\Original_price.xlsx' #Путь к прайсу с нашими ценами
path = 'c:\\Users\\Oleksiy\\Desktop\\Git\\Netology\\Excel\\' #Задали переменную с путем к рабочей папке

list_files = glob.glob(path + '*.xlsx') #Сформировали переменную, содержащую список всех файлов Excel в рабочей папке
print(glob.glob(path + '*.xlsx')) #Вывели список всех файлов с расширением Excel
for file in list_files:
    print('Обрабатываем файл: ', file)

    #Создаем страницу с именем прайса конкурента
    regxp = '\d.xlsx'
    result = re.findall(regxp, file)
    result = str(result)
    print('Тип: ', type(result))
    print('Нашли: ', result)
    sheet_name = result.replace('[','').replace('\'','').replace(']','')
    print(sheet_name)

    work_sheet = final_wb.create_sheet(sheet_name)

    #Получаем данные из нашего прайса (для сравнения)
    wb_orig = openpyxl.load_workbook(original_path)
    sheets_list_original = wb_orig.sheetnames
    sheet_active_original = wb_orig[sheets_list_original[0]]
    row_count_original = sheet_active_original.max_row
    column_count_original = sheet_active_original.max_column

    #Начинаем получать данные из прайса конкурента
    wb = openpyxl.load_workbook(file)
    sheets_list = wb.sheetnames
    sheet_active = wb[sheets_list[0]] #Сделали активной самую первую страницу
    row_count = sheet_active.max_row
    column_count = sheet_active.max_column

    print('В нашем прайсе', original_path, '\n - строк: ', row_count_original, '\n - колонок: ', column_count_original)
    print('В документе', file, '\n - строк: ', row_count, '\n - колонок: ', column_count, '\n')
    if row_count != row_count_original or column_count != column_count_original:
        print('Прайсы не совпадают, проверьте данные! \n')
        break
    else:
        print('Прайсы совпадают, работаем! \n')
        max_row = 1
        worksheet_a = 'A'
        worksheet_b = 'B'
        work_sheet_c = 'C'
        work_sheet_d = 'D'

        work_sheet['A1'] = 'Товарная позиция'
        work_sheet_a1 = work_sheet['A1']
        work_sheet_a1.font = Font(bold=True)

        work_sheet['B1'] = 'Цена у конкурента'
        work_sheet_b1 = work_sheet['B1']
        work_sheet_b1.font = Font(bold=True)

        work_sheet['C1'] = 'Цена у нас'
        work_sheet_c1 = work_sheet['C1']
        work_sheet_c1.font = Font(bold=True)

        work_sheet['D1'] = 'Вердикт'
        work_sheet_d1 = work_sheet['D1']
        work_sheet_d1.font = Font(bold=True)

        while max_row <= row_count:
            cell_a = worksheet_a + str(max_row)
            cell_b = worksheet_b + str(max_row)
            cell_c = work_sheet_c + str(max_row)
            cell_d = work_sheet_d + str(max_row)

            from_cell_a = sheet_active[cell_a].value #Получаем значение ячейки A
            from_cell_b = sheet_active[cell_b].value #Получаем значение ячейки B
            from_cell_a_orig = sheet_active_original[cell_a].value #Получаем значение ячейки А из нашего прайса
            from_cell_b_orig = sheet_active_original[cell_b].value #Получаем значение ячейки В из нашего прайса
            if from_cell_b > from_cell_b_orig:
                print(from_cell_a, ': \n - у конкурентов: ', from_cell_b, '\n - у нас: ', from_cell_b_orig)
                print('У конкурентов позиция дороже!\n')
                # Пишем в итоговый файл все
                work_sheet[cell_a] = from_cell_a
                work_sheet[cell_b] = from_cell_b
                work_sheet[cell_c] = from_cell_b_orig
                work_sheet[cell_d] = 'Позиция у конкурентов дороже!'
                work_sheet_cell_fill = work_sheet[cell_d]
                work_sheet_cell_fill.fill = PatternFill(fill_type='solid', start_color='20FF19', end_color='20FF19')
            elif from_cell_b < from_cell_b_orig:
                print(from_cell_a, ': \n - у конкурентов: ', from_cell_b, '\n - у нас: ', from_cell_b_orig)
                print('У конкурентов позиция дешевле\n')
                # Пишем в итоговый файл все
                work_sheet[cell_a] = from_cell_a
                work_sheet[cell_b] = from_cell_b
                work_sheet[cell_c] = from_cell_b_orig
                work_sheet[cell_d] = 'Позиция у конкурентов дешевле!'
                work_sheet_cell_fill = work_sheet[cell_d]
                work_sheet_cell_fill.fill = PatternFill(fill_type='solid', start_color='ff8327', end_color='ff8327')
            elif from_cell_b == from_cell_b_orig:
                print(from_cell_a, ': \n - у конкурентов: ', from_cell_b, '\n - у нас: ', from_cell_b_orig)
                print('Стоимость товара равная\n')
                #Пишем в итоговый файл все
                work_sheet[cell_a] = from_cell_a
                work_sheet[cell_b] = from_cell_b
                work_sheet[cell_c] = from_cell_b_orig
                work_sheet[cell_d] = 'Стоимость товаров равна'
                work_sheet_cell_fill_d = work_sheet[cell_d]
                work_sheet_cell_fill_d.fill = PatternFill(fill_type='solid', start_color='FFDD2D', end_color='FFDD2D')
            max_row = max_row + 1

final_wb.save('all_data.xlsx')