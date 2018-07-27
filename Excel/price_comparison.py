import openpyxl
import glob #Библиотека для работы с файлами

print('Начинаем работу: \n')

#Формируем список файлов для работы
original_path = 'c:\\Users\\Oleksiy\\Desktop\\Git\\Netology\\Excel\\Original\\Original_price.xlsx' #Путь к прайсу с нашими ценами
path = 'c:\\Users\\Oleksiy\\Desktop\\Git\\Netology\\Excel\\' #Задали переменную с путем к рабочей папке

list_files = glob.glob(path + '*.xlsx') #Сформировали переменную, содержащую список всех файлов Excel в рабочей папке
print(glob.glob(path + '*.xlsx')) #Вывели список всех файлов с расширением Excel
for file in list_files:
    print('Обрабатываем файл: ', file)

    #Получаем данные из нашего прайса (для сравнения)
    wb_orig = openpyxl.load_workbook(original_path)
    sheets_list_original = wb_orig.sheetnames
    sheet_active_original = wb_orig[sheets_list_original[0]]
    row_count_original = sheet_active_original.max_row
    column_count_original = sheet_active_original.max_column

    #Начинаем получать данные из прайса конкурента
    wb = openpyxl.load_workbook(file)
    sheets_list = wb.sheetnames
    sheet_active = wb[sheets_list[0]] #Сделали активной самую первую страницу
    row_count = sheet_active.max_row
    column_count = sheet_active.max_column

    print('В нашем прайсе', original_path, '\n - строк: ', row_count_original, '\n - колонок: ', column_count_original)
    print('В документе', file, '\n - строк: ', row_count, '\n - колонок: ', column_count, '\n')
    if row_count != row_count_original or column_count != column_count_original:
        print('Прайсы не совпадают, проверьте данные! \n')
        break
    else:
        print('Прайсы совпадают, работаем! \n')
        max_row = 1
        worksheet_a = 'A'
        worksheet_b = 'B'

        while max_row <= row_count:
            cell_a = worksheet_a + str(max_row)
            cell_b = worksheet_b + str(max_row)
            from_cell_a = sheet_active[cell_a].value #Получаем значение ячейки A #а нафига? Если прайсы одиннаковые - может проще писать в тот, что есть
            from_cell_b = sheet_active[cell_b].value #Получаем значение ячейки B
            from_cell_a_orig = sheet_active_original[cell_a].value #Получаем значение ячейки А из нашего прайса
            from_cell_b_orig = sheet_active_original[cell_b].value #Получаем значение ячейки В из нашего прайса
            if from_cell_b > from_cell_b_orig:
                print(from_cell_a, ': \n - у конкурентов: ', from_cell_b, '\n - у нас: ', from_cell_b_orig)
                print('У конкурентов позиция дороже!\n')
                # Пишем в итоговый файл все
            elif from_cell_b < from_cell_b_orig:
                print(from_cell_a, ': \n - у конкурентов: ', from_cell_b, '\n - у нас: ', from_cell_b_orig)
                print('У конкурентов позиция дешевле\n')
                # Пишем в итоговый файл все
            elif from_cell_b == from_cell_b_orig:
                print(from_cell_a, ': \n - у конкурентов: ', from_cell_b, '\n - у нас: ', from_cell_b_orig)
                print('Стоимость товара равная\n')
                #Пишем в итоговый файл все
            max_row = max_row + 1